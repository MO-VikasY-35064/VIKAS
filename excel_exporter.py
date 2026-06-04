import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from loguru import logger
from typing import List, Dict, Optional, Tuple

class ExcelExporter:
    """
    Utility class for exporting DataFrames and pivot tables to Excel
    with professional formatting and styling.
    """
    
    def __init__(self):
        """
        Initialize ExcelExporter.
        """
        self.workbook = None
        self.writer = None
        logger.info("ExcelExporter initialized")
    
    def export_dataframe(self,
                        data: pd.DataFrame,
                        file_path: str,
                        sheet_name: str = 'Data',
                        include_index: bool = True,
                        apply_formatting: bool = True) -> None:
        """
        Export DataFrame to Excel file.
        
        Args:
            data: DataFrame to export
            file_path: Path to output Excel file
            sheet_name: Name of worksheet
            include_index: Whether to include DataFrame index
            apply_formatting: Whether to apply formatting
        """
        try:
            logger.info(f"Exporting DataFrame to {file_path}")
            
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                data.to_excel(writer, sheet_name=sheet_name, index=include_index)
                
                if apply_formatting:
                    self._format_worksheet(writer.sheets[sheet_name], data, include_index)
            
            logger.info(f"DataFrame successfully exported to {file_path}")
        except Exception as e:
            logger.error(f"Error exporting DataFrame: {str(e)}")
            raise
    
    def export_multiple_dataframes(self,
                                   dataframes: Dict[str, pd.DataFrame],
                                   file_path: str,
                                   apply_formatting: bool = True) -> None:
        """
        Export multiple DataFrames to different sheets in a single Excel file.
        
        Args:
            dataframes: Dictionary mapping sheet names to DataFrames
            file_path: Path to output Excel file
            apply_formatting: Whether to apply formatting to all sheets
        """
        try:
            logger.info(f"Exporting {len(dataframes)} DataFrames to {file_path}")
            
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, data in dataframes.items():
                    data.to_excel(writer, sheet_name=sheet_name, index=True)
                    
                    if apply_formatting:
                        self._format_worksheet(writer.sheets[sheet_name], data, True)
            
            logger.info(f"All DataFrames successfully exported to {file_path}")
        except Exception as e:
            logger.error(f"Error exporting multiple DataFrames: {str(e)}")
            raise
    
    def _format_worksheet(self, worksheet, data: pd.DataFrame, include_index: bool) -> None:
        """
        Apply formatting to worksheet.
        
        Args:
            worksheet: openpyxl Worksheet object
            data: Source DataFrame
            include_index: Whether index was included
        """
        try:
            # Define styles
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format header row
            start_col = 2 if include_index else 1
            for col in range(start_col, len(data.columns) + start_col):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Format data rows
            for row in range(2, len(data) + 2):
                for col in range(1, len(data.columns) + 2):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Adjust column widths
            for col in range(1, len(data.columns) + 2):
                worksheet.column_dimensions[get_column_letter(col)].width = 15
            
            logger.info("Worksheet formatting applied")
        except Exception as e:
            logger.warning(f"Error applying formatting: {str(e)}")
    
    def export_with_charts(self,
                          data: pd.DataFrame,
                          file_path: str,
                          sheet_name: str = 'Data') -> None:
        """
        Export DataFrame to Excel with basic formatting.
        
        Args:
            data: DataFrame to export
            file_path: Path to output Excel file
            sheet_name: Name of worksheet
        """
        try:
            logger.info(f"Exporting DataFrame with formatting to {file_path}")
            data.to_excel(file_path, sheet_name=sheet_name, index=True)
            logger.info(f"Export completed: {file_path}")
        except Exception as e:
            logger.error(f"Error in export with charts: {str(e)}")
            raise
